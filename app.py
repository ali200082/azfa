#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
تطبيق ويب لإدارة بيانات الطلاب (AZFA Web)
- يعمل في المتصفح على الموبايل والحاسبة
- يقرأ/يكتب البيانات على نفس Firebase الخاص ببرنامج الحاسبة الحالي
  المسار: schools/<school_id>/students/<student_id>
- لا يعدل أي ملف من البرنامج الأصلي
"""
from flask import (Flask, render_template, request, redirect, url_for,
                   jsonify, flash, session)
import requests
import uuid
import os
import hashlib
from functools import wraps
from datetime import datetime

# ────────────── إعدادات Firebase (نفس البرنامج الحالي) ──────────────
FIREBASE_URL = "https://azfa-3ff21-default-rtdb.firebaseio.com"
FIREBASE_URL_FALLBACK = "https://falling-haze-c742.ahmed-wazir-hlail.workers.dev"
HTTP_TIMEOUT = 15

# ────────────── إعدادات المصادقة (نفس البرنامج الأصلي) ──────────────
# حساب المطور (مطابق لما في برنامج الحاسبة)
DEVELOPER_ID = "AHMEDWAZIR"
DEVELOPER_HASH = "0acb8d25fb2311d736c7ee4142ef94fc3ea3231557833dd6c664ab30699a8ed8"  # 1996124
# اسم احتياطي قديم (يقبل أيضاً)
DEVELOPER_ID_LEGACY = bytes([65, 90, 70, 65, 95, 68, 69, 86]).decode()  # AZFA_DEV
DEVELOPER_HASH_LEGACY = "69b088fc284dd65a60dd2edf0e355e6452913608f8a46b572fcf3f43decd65fb"
# حساب طوارئ افتراضي (يستخدم فقط إذا فشل الاتصال بالسحابة)
FALLBACK_USER = "AWH"
FALLBACK_PASSWORD = "1996"

# ────────────── الترجمة (العربية / الفارسية) ──────────────
TRANSLATIONS = {
    "ar": {
        "app_title": "إدارة الطلاب",
        "students_list": "قائمة الطلاب",
        "add_student": "إضافة طالب",
        "attendance": "تسجيل الحضور",
        "grades": "إدخال الدرجات",
        "reports": "التقارير",
        "exam_halls": "قاعات الامتحان",
        "users_mgmt": "إدارة المستخدمين",
        "settings": "الإعدادات",
        "support": "دعم",
        "dev_panel": "لوحة المطور",
        "switch_school": "تبديل الجامعة",
        "logout": "تسجيل الخروج",
        "login": "تسجيل الدخول",
        "username": "اسم المستخدم",
        "password": "كلمة المرور",
        "language": "اللغة",
        "arabic": "العربية",
        "persian": "الفارسية",
        "save": "حفظ",
        "present": "حاضر",
        "absent": "غائب",
        "name": "الاسم",
        "code": "الكود",
        "level": "المستوى",
        "group": "الكروب",
        "groups": "الكروبات",
        "show_students": "🔍 عرض الطلاب",
        "all_levels": "كل المستويات",
        "all_groups": "كل الكروبات",
        "prior_absences": "غياب سابق",
        "unrecorded_today": "عدد الطلاب غير المسجلين اليوم",
        "all_recorded": "تم تسجيل جميع الطلاب اليوم",
        "select_filter_hint": "اختر المستوى أو الكروب ثم اضغط عرض الطلاب",
    },
    "fa": {
        "app_title": "مدیریت دانشجویان",
        "students_list": "فهرست دانشجویان",
        "add_student": "افزودن دانشجو",
        "attendance": "ثبت حضور و غیاب",
        "grades": "ثبت نمرات",
        "reports": "گزارش ها",
        "exam_halls": "سالن های امتحان",
        "users_mgmt": "مدیریت کاربران",
        "settings": "تنظیمات",
        "support": "پشتیبانی",
        "dev_panel": "پنل توسعه دهنده",
        "switch_school": "تغییر دانشگاه",
        "logout": "خروج",
        "login": "ورود",
        "username": "نام کاربری",
        "password": "گذرواژه",
        "language": "زبان",
        "arabic": "عربی",
        "persian": "فارسی",
        "save": "ذخیره",
        "present": "حاضر",
        "absent": "غایب",
        "name": "نام",
        "code": "کد",
        "level": "سطح",
        "group": "گروه",
        "groups": "گروه ها",
        "show_students": "🔍 نمایش دانشجویان",
        "all_levels": "همه سطوح",
        "all_groups": "همه گروه ها",
        "prior_absences": "غیبت قبلی",
        "unrecorded_today": "تعداد دانشجویان ثبت نشده امروز",
        "all_recorded": "تمام دانشجویان امروز ثبت شده اند",
        "select_filter_hint": "سطح یا گروه را انتخاب کرده و دکمه نمایش را بزنید",
    },
}


def _t(key):
    lang = session.get("lang", "ar")
    return TRANSLATIONS.get(lang, TRANSLATIONS["ar"]).get(key, key)


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "azfa-web-secret-change-me-in-prod")


@app.route("/set-language/<lang>")
def set_language(lang):
    if lang in TRANSLATIONS:
        session["lang"] = lang
    return redirect(request.referrer or url_for("index"))


def _sha256(text):
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("username"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def developer_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("username"):
            return redirect(url_for("login"))
        if not session.get("is_developer"):
            flash("هذه الصفحة للمطور فقط", "error")
            return redirect(url_for("index"))
        return view(*args, **kwargs)
    return wrapped


def _is_admin():
    """المدير أو المطور."""
    return session.get("role") == "admin" or session.get("is_developer")


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("username"):
            return redirect(url_for("login"))
        if not _is_admin():
            flash("هذه الصلاحية للمدير فقط", "error")
            return redirect(url_for("index"))
        return view(*args, **kwargs)
    return wrapped


def _user_scope():
    """يرجع (assigned_level, assigned_groups) للأستاذ.
    المدير/المطور: ('' , []) أي لا قيود."""
    if _is_admin():
        return "", []
    return (
        str(session.get("assigned_level", "") or ""),
        list(session.get("assigned_groups", []) or []),
    )


def _filter_by_scope(students):
    """يقصر القائمة على المستوى/الكروبات المسموحة للأستاذ.
    المدير لا يتأثر."""
    if _is_admin():
        return students
    lvl, grps = _user_scope()
    out = students
    if lvl:
        out = [s for s in out if str(s.get("المستوى", "")) == lvl]
    if grps:
        out = [s for s in out if str(s.get("الكروب", "")) in grps]
    return out


def _can_access_student(student):
    if _is_admin():
        return True
    if not isinstance(student, dict):
        return False
    lvl, grps = _user_scope()
    if lvl and str(student.get("المستوى", "")) != lvl:
        return False
    if grps and str(student.get("الكروب", "")) not in grps:
        return False
    return True

# ────────────── حقول الطالب ──────────────
STUDENT_FIELDS = [
    ("الكود", "text"),
    ("الاسم", "text"),
    ("المستوى", "text"),
    ("الكروب", "text"),
    ("غياب", "number"),
    ("درجة الحضور", "number"),
    ("النشاط", "number"),
    ("الشفهي", "number"),
    ("التحريري", "number"),
]
NUMERIC_FIELDS = {"غياب", "درجة الحضور", "النشاط", "الشفهي", "التحريري"}


# ────────────── طبقة Firebase ──────────────
def _fb_request(method, path, **kwargs):
    """طلب HTTP مع fallback تلقائي لرابط Firebase البديل."""
    last_err = None
    for base in (FIREBASE_URL, FIREBASE_URL_FALLBACK):
        if not base:
            continue
        url = f"{base.rstrip('/')}/{path}.json"
        try:
            r = requests.request(method, url, timeout=HTTP_TIMEOUT, **kwargs)
            if r.status_code < 500:
                return r
        except Exception as e:
            last_err = e
            continue
    if last_err:
        raise last_err
    return None


def fb_get(path):
    try:
        r = _fb_request("GET", path)
        return r.json() if r is not None and r.ok else None
    except Exception as e:
        print(f"[Firebase GET error] {path}: {e}")
        return None


def fb_put(path, data):
    try:
        r = _fb_request("PUT", path, json=data)
        return r is not None and r.ok
    except Exception as e:
        print(f"[Firebase PUT error] {path}: {e}")
        return False


def fb_patch(path, data):
    try:
        r = _fb_request("PATCH", path, json=data)
        return r is not None and r.ok
    except Exception as e:
        print(f"[Firebase PATCH error] {path}: {e}")
        return False


def fb_delete(path):
    try:
        r = _fb_request("DELETE", path)
        return r is not None and r.ok
    except Exception as e:
        print(f"[Firebase DELETE error] {path}: {e}")
        return False


# ────────────── طبقة المدارس والطلاب ──────────────
def list_schools():
    """يرجع: [{id, name, count}, ...]"""
    data = fb_get("schools")
    if not isinstance(data, dict):
        return []
    schools = []
    for sid, info in data.items():
        if not isinstance(info, dict):
            continue
        schools.append({
            "id": sid,
            "name": info.get("school_name", sid),
            "count": info.get("students_count", 0),
        })
    return sorted(schools, key=lambda x: str(x["name"]))


def current_school_id():
    """يرجع معرف المدرسة المختار حالياً.
    - المستخدم العادي: محجوز بمدرسته (لا يقدر يغيرها).
    - المطور: يقدر يبدل عبر ?school=... أو من صفحة المدارس.
    """
    if session.get("is_developer"):
        sid = request.args.get("school") or session.get("school_id")
        if sid and request.args.get("school"):
            session["school_id"] = sid
        return sid
    # غير المطور: يلتزم بمدرسته فقط
    return session.get("school_id")


def students_path(school_id, student_id=None):
    if student_id:
        return f"schools/{school_id}/students/{student_id}"
    return f"schools/{school_id}/students"


def load_students(school_id):
    data = fb_get(students_path(school_id))
    if not data:
        return []
    students = []
    if isinstance(data, dict):
        for sid, s in data.items():
            if isinstance(s, dict):
                s = dict(s)
                s["id"] = sid
                students.append(s)
    elif isinstance(data, list):
        for i, s in enumerate(data):
            if isinstance(s, dict):
                s = dict(s)
                s.setdefault("id", str(i))
                students.append(s)
    return students


def get_student(school_id, sid):
    data = fb_get(students_path(school_id, sid))
    if isinstance(data, dict):
        data["id"] = sid
        return data
    return None


def _update_school_meta(school_id):
    """يحدث students_count و last_updated مثل البرنامج الأصلي."""
    students = fb_get(students_path(school_id))
    count = len(students) if isinstance(students, (dict, list)) else 0
    fb_patch(f"schools/{school_id}", {
        "students_count": count,
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated_by": "AZFA Web",
    })


def save_student(school_id, sid, payload):
    ok = fb_put(students_path(school_id, sid), payload)
    if ok:
        _update_school_meta(school_id)
    return ok


def remove_student_db(school_id, sid):
    ok = fb_delete(students_path(school_id, sid))
    if ok:
        _update_school_meta(school_id)
    return ok


def _parse_form(form):
    out = {}
    for name, _ in STUDENT_FIELDS:
        v = form.get(name, "").strip()
        if name in NUMERIC_FIELDS:
            try:
                val = float(v) if v else 0
                out[name] = int(val) if val == int(val) else val
            except ValueError:
                out[name] = 0
        else:
            out[name] = v
    return out


def _calc_total(s):
    try:
        return sum(float(s.get(f, 0) or 0) for f in
                   ("درجة الحضور", "النشاط", "الشفهي", "التحريري"))
    except (TypeError, ValueError):
        return 0


# ────────────── السياق العام للقوالب ──────────────
@app.context_processor
def inject_globals():
    schools = list_schools()
    sid = session.get("school_id")
    current = next((s for s in schools if s["id"] == sid), None)
    return {
        "all_schools": schools,
        "current_school": current,
        "current_user": session.get("username"),
        "current_role": session.get("role"),
        "is_developer": session.get("is_developer", False),
        "is_admin": _is_admin(),
        "can_modify_students": _is_admin(),
        "assigned_level": session.get("assigned_level", ""),
        "assigned_groups": session.get("assigned_groups", []),
        "lang": session.get("lang", "ar"),
        "t": _t,
        "is_demo": session.get("is_demo", False),
    }


# ────────────── المصادقة (تسجيل الدخول/الخروج) ──────────────
def cloud_authenticate(username, password):
    """يبحث في كل المدارس عن المستخدم ويتحقق من كلمة السر.
    يرجع: (school_id, role) أو (None, None)."""
    pwd_hash = _sha256(password)
    all_schools = fb_get("schools")
    if not isinstance(all_schools, dict):
        return None, None
    for sid, sdata in all_schools.items():
        if not isinstance(sdata, dict):
            continue
        cloud_users = sdata.get("users_info", {})
        if not isinstance(cloud_users, dict):
            continue
        user_info = cloud_users.get(username)
        if isinstance(user_info, dict) and user_info.get("pwd_hash") == pwd_hash:
            return sid, user_info.get("role", "teacher")
    return None, None


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if not username or not password:
            flash("الرجاء إدخال اسم المستخدم وكلمة المرور", "error")
            return render_template("login.html", username=username)

        # 1) حساب المطور (الاسم الجديد أو القديم)
        pwd_hash = _sha256(password)
        is_dev = (
            (username == DEVELOPER_ID and pwd_hash == DEVELOPER_HASH)
            or (username == DEVELOPER_ID_LEGACY and pwd_hash == DEVELOPER_HASH_LEGACY)
        )
        if is_dev:
            session.clear()
            session["username"] = username
            session["role"] = "admin"
            session["is_developer"] = True
            flash(f"مرحباً أيها المطور 👋", "success")
            return redirect(request.args.get("next") or url_for("schools_page"))

        # 2) حسابات السحابة (من schools/<id>/users_info)
        school_id, role = cloud_authenticate(username, password)
        if school_id:
            # نقرأ صلاحيات الأستاذ (مستوى/كروبات معينة)
            user_info = fb_get(f"schools/{school_id}/users_info/{username}") or {}
            session.clear()
            session["username"] = username
            session["role"] = role
            session["is_developer"] = False
            session["school_id"] = school_id
            session["assigned_level"] = str(user_info.get("assigned_level", "") or "")
            ag = user_info.get("assigned_groups", [])
            session["assigned_groups"] = list(ag) if isinstance(ag, list) else []
            flash(f"مرحباً {username} 👋", "success")
            return redirect(request.args.get("next") or url_for("index"))

        # 3) حساب طوارئ افتراضي
        if username == FALLBACK_USER and password == FALLBACK_PASSWORD:
            session.clear()
            session["username"] = username
            session["role"] = "admin"
            session["is_developer"] = False
            flash("تم الدخول بالحساب الافتراضي — اختر المدرسة", "success")
            return redirect(url_for("schools_page"))

        flash("اسم المستخدم أو كلمة المرور غير صحيحة", "error")
        return render_template("login.html", username=username)

    if session.get("username"):
        return redirect(url_for("index"))
    return render_template("login.html", username="")


@app.route("/logout")
def logout():
    user = session.get("username", "")
    session.clear()
    flash(f"تم تسجيل الخروج ({user})", "success")
    return redirect(url_for("login"))


# ────────────── المسارات ──────────────
@app.route("/")
@login_required
def index():
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))

    q = request.args.get("q", "").strip()
    level = request.args.get("level", "").strip()
    group = request.args.get("group", "").strip()

    all_students = load_students(school_id)
    all_students = _filter_by_scope(all_students)
    students = all_students

    if q:
        ql = q.lower()
        students = [s for s in students
                    if ql in str(s.get("الاسم", "")).lower()
                    or ql in str(s.get("الكود", "")).lower()]
    if level:
        students = [s for s in students if str(s.get("المستوى", "")) == level]
    if group:
        students = [s for s in students if str(s.get("الكروب", "")) == group]

    for s in students:
        s["_total"] = _calc_total(s)

    levels = sorted({str(s.get("المستوى", "")) for s in all_students if s.get("المستوى")})
    groups = sorted({str(s.get("الكروب", "")) for s in all_students if s.get("الكروب")})

    return render_template("index.html",
                           students=students, q=q, level=level, group=group,
                           levels=levels, groups=groups,
                           fields=STUDENT_FIELDS,
                           total_count=len(all_students))


@app.route("/schools")
@login_required
def schools_page():
    return render_template("schools.html", schools=list_schools())


@app.route("/schools/select/<school_id>")
@login_required
def select_school(school_id):
    if not session.get("is_developer"):
        flash("لا يمكنك تغيير المدرسة. مدرستك محددة من قبل الإدارة.", "error")
        return redirect(url_for("index"))
    session["school_id"] = school_id
    flash("تم اختيار المدرسة ✓", "success")
    return redirect(url_for("index"))


@app.route("/student/new", methods=["GET", "POST"])
@login_required
@admin_required
def new_student():
    school_id = current_school_id()
    if not school_id:
        flash("اختر مدرسة أولاً", "error")
        return redirect(url_for("schools_page"))
    if request.method == "POST":
        data = _parse_form(request.form)
        if not data.get("الاسم"):
            flash("الاسم مطلوب", "error")
            return render_template("form.html", student=data, fields=STUDENT_FIELDS, action="إضافة")
        sid = data.get("الكود", "").strip() or str(uuid.uuid4())[:12]
        if get_student(school_id, sid):
            sid = str(uuid.uuid4())[:12]
        if save_student(school_id, sid, data):
            flash("تمت الإضافة بنجاح ✓", "success")
            return redirect(url_for("index"))
        flash("فشل الحفظ — تحقق من الاتصال بالإنترنت", "error")
        return render_template("form.html", student=data, fields=STUDENT_FIELDS, action="إضافة")
    return render_template("form.html", student={}, fields=STUDENT_FIELDS, action="إضافة")


@app.route("/student/<sid>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def edit_student(sid):
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))
    student = get_student(school_id, sid)
    if not student:
        flash("الطالب غير موجود", "error")
        return redirect(url_for("index"))
    if request.method == "POST":
        data = _parse_form(request.form)
        if save_student(school_id, sid, data):
            flash("تم التعديل ✓", "success")
            return redirect(url_for("index"))
        flash("فشل الحفظ", "error")
    return render_template("form.html", student=student, fields=STUDENT_FIELDS, action="تعديل")


@app.route("/student/<sid>/delete", methods=["POST"])
@login_required
@admin_required
def remove_student(sid):
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))
    if remove_student_db(school_id, sid):
        flash("تم الحذف ✓", "success")
    else:
        flash("فشل الحذف", "error")
    return redirect(url_for("index"))


@app.route("/student/<sid>")
@login_required
def view_student(sid):
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))
    student = get_student(school_id, sid)
    if not student:
        flash("الطالب غير موجود", "error")
        return redirect(url_for("index"))
    if not _can_access_student(student):
        flash("لا تملك صلاحية لعرض هذا الطالب", "error")
        return redirect(url_for("index"))
    student["_total"] = _calc_total(student)
    return render_template("view.html", student=student, fields=STUDENT_FIELDS)


@app.route("/api/students")
@login_required
def api_students():
    school_id = current_school_id()
    if not school_id:
        return jsonify([])
    return jsonify(_filter_by_scope(load_students(school_id)))


@app.route("/api/schools")
@login_required
def api_schools():
    return jsonify(list_schools())


@app.route("/api/groups")
@login_required
def api_groups():
    """يرجع الكروبات المتوفرة في مستوى معيّن (لتعبئة القوائم ديناميكياً)."""
    school_id = current_school_id()
    if not school_id:
        return jsonify([])
    level = request.args.get("level", "").strip()
    students_all = load_students(school_id)
    groups = set()
    for s in students_all:
        if not isinstance(s, dict):
            continue
        if level and str(s.get("المستوى", "")) != level:
            continue
        g = str(s.get("الكروب", "")).strip()
        if g:
            groups.add(g)
    return jsonify(sorted(groups))


@app.route("/health")
def health():
    schools = list_schools()
    return jsonify({
        "firebase": "ok" if schools else "fail",
        "schools_count": len(schools),
        "time": datetime.now().isoformat(),
    })


# ════════════════════════════════════════════════════════════════════
# 🟢 تسجيل الحضور والغياب
# ════════════════════════════════════════════════════════════════════
def _attendance_score(absences):
    """نفس صيغة البرنامج: 20 - (الغياب / 2)، بحد أدنى 0"""
    try:
        return max(0, 20 - (float(absences or 0) / 2))
    except (TypeError, ValueError):
        return 20


@app.route("/attendance", methods=["GET", "POST"])
@login_required
def attendance_page():
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))

    all_students = load_students(school_id)
    all_students = _filter_by_scope(all_students)
    levels = sorted({str(s.get("المستوى", "")) for s in all_students if s.get("المستوى")})
    groups = sorted({str(s.get("الكروب", "")) for s in all_students if s.get("الكروب")})

    # خريطة الكروبات لكل مستوى (لفلترة الكروبات حسب المستوى المختار)
    groups_by_level = {}
    for s in all_students:
        lvl = str(s.get("المستوى", ""))
        grp = str(s.get("الكروب", ""))
        if not lvl or not grp:
            continue
        groups_by_level.setdefault(lvl, set()).add(grp)
    groups_by_level = {k: sorted(v) for k, v in groups_by_level.items()}

    sel_level = request.args.get("level", "").strip() or request.form.get("level", "").strip()
    sel_groups = request.values.getlist("groups")
    sel_groups = [g for g in sel_groups if g]
    today = datetime.now().strftime("%Y-%m-%d")

    # تصفية الطلاب حسب المستوى والكروبات (عدة)
    filtered = all_students
    if sel_level:
        filtered = [s for s in filtered if str(s.get("المستوى", "")) == sel_level]
    if sel_groups:
        filtered = [s for s in filtered if str(s.get("الكروب", "")) in sel_groups]

    # POST: تسجيل الحضور
    if request.method == "POST":
        action = request.form.get("action", "")  # present | absent
        student_ids = request.form.getlist("students")
        if not student_ids:
            flash("لم تحدد أي طالب", "error")
        else:
            success = 0
            already = 0
            for sid in student_ids:
                stu = get_student(school_id, sid)
                if not stu:
                    continue
                if not _can_access_student(stu):
                    continue
                log = stu.get("سجل_الحضور", [])
                if not isinstance(log, list):
                    log = []
                # تحقق إذا تم تسجيله اليوم
                already_today = any(
                    isinstance(e, dict) and str(e.get("التاريخ", "")).startswith(today)
                    for e in log
                )
                if already_today:
                    already += 1
                    continue
                log.append({
                    "التاريخ": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "الحالة": "حاضر" if action == "present" else "غائب",
                })
                stu["سجل_الحضور"] = log
                if action == "absent":
                    stu["غياب"] = int(stu.get("غياب", 0) or 0) + 1
                stu["درجة الحضور"] = _attendance_score(stu.get("غياب", 0))
                stu.pop("id", None)
                save_student(school_id, sid, stu)
                success += 1
            if success:
                flash(f"تم تسجيل {success} طالب ✓" + (f" (تم تجاهل {already} مسجلين مسبقاً)" if already else ""), "success")
            elif already:
                flash(f"كل الطلاب المحددين مسجلين اليوم بالفعل", "error")

        # تحديث القائمة بعد التسجيل
        all_students = load_students(school_id)
        all_students = _filter_by_scope(all_students)
        filtered = all_students
        if sel_level:
            filtered = [s for s in filtered if str(s.get("المستوى", "")) == sel_level]
        if sel_groups:
            filtered = [s for s in filtered if str(s.get("الكروب", "")) in sel_groups]

    # الطلاب غير المسجلين اليوم
    unrecorded = []
    for s in filtered:
        log = s.get("سجل_الحضور", [])
        if not isinstance(log, list):
            log = []
        is_today = any(
            isinstance(e, dict) and str(e.get("التاريخ", "")).startswith(today)
            for e in log
        )
        if not is_today:
            unrecorded.append(s)

    return render_template("attendance.html",
                           students=sorted(unrecorded, key=lambda s: str(s.get("الكود", ""))),
                           levels=levels, groups=groups,
                           groups_by_level=groups_by_level,
                           sel_level=sel_level, sel_groups=sel_groups,
                           today=today)


# ════════════════════════════════════════════════════════════════════
# 🟢 إدخال الدرجات
# ════════════════════════════════════════════════════════════════════
@app.route("/grades", methods=["GET", "POST"])
@login_required
def grades_page():
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))

    sid_q = request.args.get("sid", "").strip()
    code_q = request.args.get("code", "").strip()
    selected = None

    # البحث بالكود
    if code_q and not sid_q:
        for s in load_students(school_id):
            if str(s.get("الكود", "")) == code_q:
                selected = s
                sid_q = s.get("id")
                break
        if not selected:
            flash(f"لا يوجد طالب بالكود {code_q}", "error")

    if sid_q and not selected:
        selected = get_student(school_id, sid_q)

    if request.method == "POST" and selected:
        # الأستاذ يدخل النشاط فقط — المدير/المطور يدخل كل الدرجات
        try:
            nashat = float(request.form.get("النشاط", selected.get("النشاط", 0)) or 0)
            if _is_admin():
                shafahi = float(request.form.get("الشفهي", selected.get("الشفهي", 0)) or 0)
                tahriri = float(request.form.get("التحريري", selected.get("التحريري", 0)) or 0)
            else:
                shafahi = float(selected.get("الشفهي", 0) or 0)
                tahriri = float(selected.get("التحريري", 0) or 0)
        except ValueError:
            flash("الدرجات يجب أن تكون أرقاماً", "error")
            return render_template("grades.html", student=selected)

        # تحقق من الحدود (مثل البرنامج)
        if not (0 <= nashat <= 10):
            flash("النشاط يجب أن يكون بين 0 و 10", "error")
        elif not (0 <= shafahi <= 20):
            flash("الشفهي يجب أن يكون بين 0 و 20", "error")
        elif not (0 <= tahriri <= 60):
            flash("التحريري يجب أن يكون بين 0 و 60", "error")
        else:
            selected["النشاط"] = nashat
            selected["الشفهي"] = shafahi
            selected["التحريري"] = tahriri
            selected["درجة الحضور"] = _attendance_score(selected.get("غياب", 0))
            sid_actual = selected.pop("id", sid_q)
            if save_student(school_id, sid_actual, selected):
                flash(f"تم حفظ درجات {selected.get('الاسم', '')} ✓", "success")
                return redirect(url_for("grades_page", sid=sid_actual))
            flash("فشل الحفظ", "error")

    if selected:
        selected["_total"] = _calc_total(selected)
    return render_template("grades.html", student=selected, code_q=code_q)


# ════════════════════════════════════════════════════════════════════
# 🟢 التقارير + تصدير Excel
# ════════════════════════════════════════════════════════════════════
@app.route("/reports")
@login_required
def reports_page():
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))

    q = request.args.get("q", "").strip()
    level = request.args.get("level", "").strip()
    group = request.args.get("group", "").strip()

    all_students = load_students(school_id)
    all_students = _filter_by_scope(all_students)
    students = list(all_students)
    if q:
        ql = q.lower()
        students = [s for s in students if ql in str(s.get("الاسم", "")).lower()]
    if level:
        students = [s for s in students if str(s.get("المستوى", "")) == level]
    if group:
        students = [s for s in students if str(s.get("الكروب", "")) == group]

    # ترتيب: مستوى ← كروب ← كود
    def _sort_key(s):
        return (str(s.get("المستوى", "")), str(s.get("الكروب", "")), str(s.get("الكود", "")))
    students.sort(key=_sort_key)

    for s in students:
        s["_total"] = _calc_total(s)
        s["_result"] = "ناجح" if s["_total"] >= 60 else "راسب"

    levels = sorted({str(s.get("المستوى", "")) for s in all_students if s.get("المستوى")})
    groups = sorted({str(s.get("الكروب", "")) for s in all_students if s.get("الكروب")})

    return render_template("reports.html",
                           students=students, q=q, level=level, group=group,
                           levels=levels, groups=groups)


@app.route("/reports/export")
@login_required
def reports_export():
    """تصدير Excel للتقرير الحالي (يقبل نفس فلاتر /reports)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
    except ImportError:
        flash("مكتبة Excel غير متوفرة على السيرفر", "error")
        return redirect(url_for("reports_page"))

    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))

    q = request.args.get("q", "").strip()
    level = request.args.get("level", "").strip()
    group = request.args.get("group", "").strip()

    students = load_students(school_id)
    students = _filter_by_scope(students)
    if q:
        ql = q.lower()
        students = [s for s in students if ql in str(s.get("الاسم", "")).lower()]
    if level:
        students = [s for s in students if str(s.get("المستوى", "")) == level]
    if group:
        students = [s for s in students if str(s.get("الكروب", "")) == group]
    students.sort(key=lambda s: (str(s.get("المستوى", "")), str(s.get("الكروب", "")), str(s.get("الكود", ""))))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "التقرير"
    ws.sheet_view.rightToLeft = True

    headers = ["الكود", "الاسم", "المستوى", "الكروب",
               "درجة الحضور", "النشاط", "الشفهي", "التحريري", "المجموع", "النتيجة"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="0d6efd")
    head_font = Font(color="FFFFFF", bold=True, size=12)
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for i, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.fill = head_fill
        c.font = head_font
        c.alignment = center
        c.border = border

    for r, s in enumerate(students, 2):
        total = _calc_total(s)
        result = "ناجح" if total >= 60 else "راسب"
        row = [s.get("الكود", ""), s.get("الاسم", ""), s.get("المستوى", ""),
               s.get("الكروب", ""), s.get("درجة الحضور", 0),
               s.get("النشاط", 0), s.get("الشفهي", 0), s.get("التحريري", 0),
               total, result]
        for i, val in enumerate(row, 1):
            c = ws.cell(row=r, column=i, value=val)
            c.alignment = center
            c.border = border
            if i == 10:  # النتيجة
                c.fill = PatternFill("solid", fgColor="d1e7dd" if result == "ناجح" else "f8d7da")

    # عرض الأعمدة
    widths = [10, 25, 12, 10, 12, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/reports/student/<sid>")
@login_required
def student_report(sid):
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))
    student = get_student(school_id, sid)
    if not student:
        flash("الطالب غير موجود", "error")
        return redirect(url_for("reports_page"))
    if not _can_access_student(student):
        flash("لا تملك صلاحية لعرض هذا الطالب", "error")
        return redirect(url_for("reports_page"))

    student["_total"] = _calc_total(student)
    student["_result"] = "ناجح" if student["_total"] >= 60 else "راسب"
    log = student.get("سجل_الحضور", [])
    if not isinstance(log, list):
        log = []
    present_days = sum(1 for e in log if isinstance(e, dict) and e.get("الحالة") == "حاضر")
    absent_days = sum(1 for e in log if isinstance(e, dict) and e.get("الحالة") == "غائب")
    return render_template("student_report.html", student=student,
                           log=list(reversed(log)),
                           present_days=present_days, absent_days=absent_days)


# ═══════════════════════════════════════════════════════════════════
# 🟢 تقرير النشر (الدرجات النهائية) — للطباعة
# ═══════════════════════════════════════════════════════════════════
@app.route("/reports/final")
@login_required
def final_grades():
    """تقرير الدرجات النهائية (للنشر والطباعة)."""
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))

    level = request.args.get("level", "").strip()
    group = request.args.get("group", "").strip()

    all_students = _filter_by_scope(load_students(school_id))
    students = list(all_students)
    if level:
        students = [s for s in students if str(s.get("المستوى", "")) == level]
    if group:
        students = [s for s in students if str(s.get("الكروب", "")) == group]
    students.sort(key=lambda s: (str(s.get("المستوى", "")), str(s.get("الكروب", "")), str(s.get("الكود", ""))))

    for s in students:
        s["_total"] = _calc_total(s)
        s["_result"] = "ناجح" if s["_total"] >= 60 else "راسب"

    pass_count = sum(1 for s in students if s["_result"] == "ناجح")
    fail_count = len(students) - pass_count

    levels = sorted({str(s.get("المستوى", "")) for s in all_students if s.get("المستوى")})
    groups = sorted({str(s.get("الكروب", "")) for s in all_students if s.get("الكروب")})

    return render_template("final_grades.html",
                           students=students, level=level, group=group,
                           levels=levels, groups=groups,
                           pass_count=pass_count, fail_count=fail_count,
                           today=datetime.now().strftime("%Y-%m-%d"))


# ═══════════════════════════════════════════════════════════════════
# 🟢 تقرير إحصائي (عدد الطلاب / الناجحين / الراسبين حسب المستوى)
# ═══════════════════════════════════════════════════════════════════
@app.route("/reports/statistics")
@login_required
def statistics_report():
    """تقرير إحصائي عام على مستوى المدرسة."""
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))

    students_all = _filter_by_scope(load_students(school_id))
    total = len(students_all)
    pass_count = 0
    fail_count = 0
    sum_total = 0.0
    by_level = {}
    by_group = {}
    total_absences = 0

    rows = []
    for s in students_all:
        t = _calc_total(s)
        sum_total += t
        result = "ناجح" if t >= 60 else "راسب"
        if result == "ناجح":
            pass_count += 1
        else:
            fail_count += 1
        try:
            total_absences += int(float(s.get("غياب", 0) or 0))
        except (TypeError, ValueError):
            pass
        lvl = str(s.get("المستوى", "غير محدد"))
        grp = str(s.get("الكروب", "غير محدد"))
        by_level.setdefault(lvl, {"total": 0, "pass": 0, "fail": 0})
        by_level[lvl]["total"] += 1
        by_level[lvl]["pass" if result == "ناجح" else "fail"] += 1
        by_group.setdefault(grp, {"total": 0, "pass": 0, "fail": 0})
        by_group[grp]["total"] += 1
        by_group[grp]["pass" if result == "ناجح" else "fail"] += 1
        s_copy = dict(s)
        s_copy["_total"] = t
        s_copy["_result"] = result
        rows.append(s_copy)

    rows.sort(key=lambda s: (str(s.get("المستوى", "")), str(s.get("الكروب", "")), str(s.get("الكود", ""))))

    avg = (sum_total / total) if total else 0
    pass_rate = (pass_count / total * 100) if total else 0

    return render_template("statistics.html",
                           students=rows,
                           total=total, pass_count=pass_count, fail_count=fail_count,
                           avg=round(avg, 2), pass_rate=round(pass_rate, 1),
                           total_absences=total_absences,
                           by_level=sorted(by_level.items()),
                           by_group=sorted(by_group.items()))


# ═══════════════════════════════════════════════════════════════════
# 🟢 تقرير الغياب
# ═══════════════════════════════════════════════════════════════════
@app.route("/reports/absence")
@login_required
def absence_report():
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))
    students = _filter_by_scope(load_students(school_id))
    rows = []
    for s in students:
        try:
            a = int(float(s.get("غياب", 0) or 0))
        except (TypeError, ValueError):
            a = 0
        if a > 0:
            s["_absences"] = a
            rows.append(s)
    rows.sort(key=lambda s: s["_absences"], reverse=True)
    return render_template("absence_report.html", students=rows)


# ════════════════════════════════════════════════════════════════════
# 🟢 توزيع قاعات الامتحان
# ════════════════════════════════════════════════════════════════════
@app.route("/exam-halls", methods=["GET", "POST"])
@login_required
def exam_halls():
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))

    all_students = load_students(school_id)
    all_students = _filter_by_scope(all_students)
    levels = sorted({str(s.get("المستوى", "")) for s in all_students if s.get("المستوى")})
    groups = sorted({str(s.get("الكروب", "")) for s in all_students if s.get("الكروب")})

    sel_levels = request.values.getlist("levels")
    sel_groups = request.values.getlist("groups")
    pages = []

    if sel_levels or sel_groups:
        chosen = all_students
        if sel_levels:
            chosen = [s for s in chosen if str(s.get("المستوى", "")) in sel_levels]
        if sel_groups:
            chosen = [s for s in chosen if str(s.get("الكروب", "")) in sel_groups]

        # خلط أو ترتيب
        if len(sel_groups) > 1 or len(sel_levels) > 1:
            import random
            random.shuffle(chosen)
        else:
            chosen.sort(key=lambda s: str(s.get("الكود", "")), reverse=True)

        # 18 طالب لكل صفحة
        PER_PAGE = 18
        for i in range(0, len(chosen), PER_PAGE):
            pages.append(chosen[i:i + PER_PAGE])

    return render_template("exam_halls.html",
                           levels=levels, groups=groups,
                           groups_by_level=groups_by_level,
                           sel_levels=sel_levels, sel_groups=sel_groups,
                           pages=pages,
                           total=sum(len(p) for p in pages))


# ════════════════════════════════════════════════════════════════════
# 🟢 إدارة الأساتذة (مدير المدرسة فقط)
# ════════════════════════════════════════════════════════════════════
@app.route("/teachers", methods=["GET", "POST"])
@login_required
@admin_required
def teachers_page():
    school_id = current_school_id()
    if not school_id:
        return redirect(url_for("schools_page"))

    if request.method == "POST":
        action = request.form.get("action")
        username = request.form.get("username", "").strip()
        if action == "add":
            password = request.form.get("password", "")
            role = request.form.get("role", "teacher")
            assigned_level = request.form.get("assigned_level", "").strip()
            # دعم الإرسال المتعدد من <select multiple> أو نص مفصول بفواصل
            ag_list = request.form.getlist("assigned_groups")
            if len(ag_list) == 1 and "," in ag_list[0]:
                ag_list = ag_list[0].split(",")
            assigned_groups = [g.strip() for g in ag_list if g and g.strip()]
            if not username or not password:
                flash("اسم المستخدم وكلمة المرور مطلوبان", "error")
            else:
                fb_put(f"schools/{school_id}/users_info/{username}", {
                    "pwd_hash": _sha256(password),
                    "pwd_plain": password,
                    "role": role,
                    "assigned_level": assigned_level,
                    "assigned_groups": assigned_groups,
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "created_by": session.get("username"),
                })
                flash(f"تمت إضافة المستخدم {username} ✓", "success")
        elif action == "delete":
            if username == session.get("username"):
                flash("لا يمكنك حذف حسابك الحالي", "error")
            else:
                fb_delete(f"schools/{school_id}/users_info/{username}")
                flash(f"تم حذف {username} ✓", "success")
        elif action == "reset_pass":
            new_pass = request.form.get("new_password", "")
            if username and new_pass:
                fb_patch(f"schools/{school_id}/users_info/{username}", {
                    "pwd_hash": _sha256(new_pass),
                    "pwd_plain": new_pass,
                })
                flash(f"تم تحديث كلمة مرور {username} ✓", "success")
        return redirect(url_for("teachers_page"))

    users_data = fb_get(f"schools/{school_id}/users_info") or {}
    users = []
    if isinstance(users_data, dict):
        for u, info in users_data.items():
            if isinstance(info, dict):
                users.append({
                    "username": u,
                    "role": info.get("role", "teacher"),
                    "assigned_level": info.get("assigned_level", ""),
                    "assigned_groups": info.get("assigned_groups", []),
                    "created_at": info.get("created_at", ""),
                    "pwd_plain": info.get("pwd_plain", "") if session.get("is_developer") else "",
                })
    users.sort(key=lambda x: x["username"])

    all_students = load_students(school_id)
    levels = sorted({str(s.get("المستوى", "")) for s in all_students if s.get("المستوى")})
    return render_template("teachers.html", users=users, levels=levels)


# ════════════════════════════════════════════════════════════════════
# 🟢 إعدادات الحساب (تغيير كلمة المرور)
# ════════════════════════════════════════════════════════════════════
@app.route("/account", methods=["GET", "POST"])
@login_required
def account_page():
    if request.method == "POST":
        old = request.form.get("old_password", "")
        new = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        if not new or new != confirm:
            flash("كلمة المرور الجديدة غير متطابقة", "error")
        elif len(new) < 4:
            flash("كلمة المرور قصيرة جداً (4 أحرف فأكثر)", "error")
        else:
            username = session.get("username")
            sid = session.get("school_id")
            if session.get("is_developer"):
                flash("لا يمكن تغيير كلمة المطور من الويب — استخدم البرنامج الأصلي", "error")
            elif sid and username:
                user_info = fb_get(f"schools/{sid}/users_info/{username}")
                if isinstance(user_info, dict) and user_info.get("pwd_hash") == _sha256(old):
                    fb_patch(f"schools/{sid}/users_info/{username}", {"pwd_hash": _sha256(new)})
                    flash("تم تغيير كلمة المرور ✓", "success")
                else:
                    flash("كلمة المرور الحالية غير صحيحة", "error")
            else:
                flash("لا يمكن تغيير كلمة المرور لهذا الحساب", "error")
    return render_template("account.html")


# ════════════════════════════════════════════════════════════════════
# 🟢 صفحة الدعم
# ════════════════════════════════════════════════════════════════════
@app.route("/support")
@login_required
def support_page():
    return render_template("support.html")


# ════════════════════════════════════════════════════════════════════
# 🟢 لوحة المطور (إدارة الجامعات/المدارس)
# ════════════════════════════════════════════════════════════════════
@app.route("/dev/dashboard", methods=["GET", "POST"])
@developer_required
def dev_dashboard():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add_school":
            school_id = request.form.get("school_id", "").strip()
            school_name = request.form.get("school_name", "").strip()
            if not school_id or not school_name:
                flash("الرجاء إدخال معرّف واسم الجامعة", "error")
            else:
                fb_put(f"schools/{school_id}", {
                    "school_name": school_name,
                    "students_count": 0,
                    "active": True,
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "created_by": session.get("username"),
                })
                flash(f"تمت إضافة الجامعة {school_name} ✓", "success")
        elif action == "delete_school":
            school_id = request.form.get("school_id", "").strip()
            if school_id:
                fb_delete(f"schools/{school_id}")
                flash(f"تم حذف الجامعة {school_id} ✓", "success")
        elif action == "toggle_school":
            school_id = request.form.get("school_id", "").strip()
            active = request.form.get("active") == "1"
            fb_patch(f"schools/{school_id}", {"active": active})
            flash("تم تحديث الحالة ✓", "success")
        return redirect(url_for("dev_dashboard"))

    # جمع التفاصيل
    data = fb_get("schools") or {}
    schools = []
    if isinstance(data, dict):
        for sid, info in data.items():
            if not isinstance(info, dict):
                continue
            users = info.get("users_info", {}) or {}
            schools.append({
                "id": sid,
                "name": info.get("school_name", sid),
                "count": info.get("students_count", 0),
                "active": info.get("active", True),
                "last_updated": info.get("last_updated", ""),
                "users_count": len(users) if isinstance(users, dict) else 0,
            })
    schools.sort(key=lambda x: str(x["name"]))
    return render_template("dev_dashboard.html", schools=schools)


@app.route("/dev/users")
@developer_required
def dev_users():
    data = fb_get("schools") or {}
    rows = []
    if isinstance(data, dict):
        for sid, info in data.items():
            if not isinstance(info, dict):
                continue
            school_name = info.get("school_name", sid)
            users = info.get("users_info", {}) or {}
            if not isinstance(users, dict):
                continue
            for uname, uinfo in users.items():
                if not isinstance(uinfo, dict):
                    continue
                rows.append({
                    "school_id": sid,
                    "school_name": school_name,
                    "username": uname,
                    "role": uinfo.get("role", "teacher"),
                    "password": uinfo.get("pwd_plain", ""),
                    "has_plain": bool(uinfo.get("pwd_plain")),
                    "assigned_level": uinfo.get("assigned_level", ""),
                    "assigned_groups": uinfo.get("assigned_groups", []),
                    "created_at": uinfo.get("created_at", ""),
                })
    rows.sort(key=lambda r: (str(r["school_name"]), str(r["username"])))
    return render_template("dev_users.html", rows=rows)


# ════════════════════════════════════════════════════════════════════
# 🟢 رابط الديمو المؤقت (يتحكّم به المطوّر فقط)
# ════════════════════════════════════════════════════════════════════
import secrets


def _get_demo_config():
    cfg = fb_get("demo_access") or {}
    if not isinstance(cfg, dict):
        cfg = {}
    return cfg


@app.route("/dev/demo", methods=["GET", "POST"])
@developer_required
def dev_demo():
    cfg = _get_demo_config()
    schools_list = list_schools()

    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "enable":
            token = secrets.token_urlsafe(12)
            role = request.form.get("role", "teacher")
            school_id = request.form.get("school_id", "").strip()
            note = request.form.get("note", "").strip()
            cfg = {
                "enabled": True,
                "token": token,
                "role": role if role in ("admin", "teacher") else "teacher",
                "school_id": school_id,
                "note": note,
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "created_by": session.get("username", ""),
            }
            fb_put("demo_access", cfg)
            flash("✓ تم إنشاء رابط الديمو", "success")
        elif action == "disable":
            fb_put("demo_access", {"enabled": False})
            flash("✓ تم إيقاف رابط الديمو — الرابط الأساسي يعمل كالمعتاد", "success")
        elif action == "delete":
            fb_delete("demo_access")
            flash("✓ تم حذف رابط الديمو نهائياً", "success")
        return redirect(url_for("dev_demo"))

    demo_url = ""
    if cfg.get("enabled") and cfg.get("token"):
        demo_url = url_for("demo_access", token=cfg["token"], _external=True)
    return render_template("dev_demo.html", cfg=cfg, demo_url=demo_url, schools=schools_list)


@app.route("/demo/<token>")
def demo_access(token):
    """دخول مؤقت عبر رابط الديمو — لا يحتاج اسم مستخدم/كلمة مرور."""
    cfg = _get_demo_config()
    if not cfg.get("enabled") or cfg.get("token") != token:
        return render_template("demo_expired.html"), 403

    role = cfg.get("role", "teacher")
    school_id = cfg.get("school_id", "")

    session.clear()
    session["username"] = "ضيف_تجريبي"
    session["role"] = role
    session["is_developer"] = False
    session["is_demo"] = True
    if school_id:
        session["school_id"] = school_id
    session["assigned_level"] = ""
    session["assigned_groups"] = []
    flash("👋 مرحباً بك في النسخة التجريبية — هذا حساب ضيف مؤقت", "success")
    if school_id:
        return redirect(url_for("index"))
    return redirect(url_for("schools_page"))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"\n🌐 افتح المتصفح على: http://localhost:{port}")
    print(f"📱 من الموبايل (نفس الواي فاي): http://<IP-الحاسبة>:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=False)
